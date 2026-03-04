"""
Main pipeline for mapping institutions from CGSpace → CLARISA
Reads an Excel file, searches for each institution in CLARISA using hybrid search
(vector embeddings + RapidFuzz) and generates a report with the results
"""
import os
import pandas as pd
from tqdm import tqdm
from rapidfuzz import fuzz
from dotenv import load_dotenv
from openpyxl.styles import Alignment

# Import project modules
from logger.logger_util import get_logger
from src.utils import clean_text_for_matching
from src.web_search import search_institution_online
from src.embeddings import get_embedding, embedding_to_list, cosine_similarity
from src.supabase_client import search_by_name_embedding, search_combined, count_institutions


load_dotenv()
logger = get_logger()

# ============================================================================
# CONFIGURATION
# ============================================================================
THRESHOLD_EMBEDDINGS = 0.2  # Threshold for vector search (lower = more candidates)
THRESHOLD_FINAL = 0.6       # Threshold to consider a valid match (final score)
NAME_WEIGHT = 0.7           # Weight of name in combined search (0-1)
ACRONYM_WEIGHT = 0.3        # Weight of acronym in combined search (0-1)
COSINE_WEIGHT = 0.50        # Weight of cosine similarity in final score
FUZZ_NAME_WEIGHT = 0.40     # Weight of RapidFuzz name in final score
FUZZ_ACRONYM_WEIGHT = 0.10  # Weight of RapidFuzz acronym in final score
ENABLE_WEB_SEARCH = True    # Enable web search fallback when no match in CLARISA

# Excel cell character limit
EXCEL_CELL_CHAR_LIMIT = 32767

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def sanitize_for_excel(text: str) -> str:
    """
    Sanitize text for Excel compatibility - only truncate if needed
    
    Args:
        text: Text to sanitize
        
    Returns:
        str: Text safe for Excel (truncated if exceeds limit)
    """
    if not text:
        return ''
    
    text = str(text)
    
    # Only truncate if exceeds Excel's character limit
    if len(text) > EXCEL_CELL_CHAR_LIMIT:
        logger.warning(f"Text truncated from {len(text)} to {EXCEL_CELL_CHAR_LIMIT} chars")
        text = text[:EXCEL_CELL_CHAR_LIMIT - 50] + "\n\n[... TRUNCATED ...]"
    
    return text


# ============================================================================
# HYBRID SEARCH FUNCTION
# ============================================================================

def search_institution_for_excel(partner_name, acronym=None):
    """
    Search for an institution using hybrid search (embeddings + RapidFuzz)
    Optimized for batch processing from Excel
    
    Args:
        partner_name: Institution name
        acronym: Optional acronym
        
    Returns:
        dict: Result with best match and scores, or None if no match
    """
    # Validate input
    if not partner_name or str(partner_name).lower() in ['nan', 'none', '']:
        return None
    
    partner_name = str(partner_name).strip()
    acronym = str(acronym).strip() if acronym and str(acronym).lower() not in ['nan', 'none', ''] else None
    
    # STEP 1: Generate query embeddings
    name_embedding = get_embedding(partner_name)
    name_embedding_list = embedding_to_list(name_embedding)
    
    # STEP 2: Vector search in Supabase (Top 5)
    if acronym:
        # Combined search (name + acronym)
        acronym_embedding = get_embedding(acronym)
        acronym_embedding_list = embedding_to_list(acronym_embedding)
        
        candidates = search_combined(
            name_embedding=name_embedding_list,
            acronym_embedding=acronym_embedding_list,
            name_weight=NAME_WEIGHT,
            acronym_weight=ACRONYM_WEIGHT,
            threshold=THRESHOLD_EMBEDDINGS,
            limit=5
        )
    else:
        # By name only
        candidates = search_by_name_embedding(
            query_embedding=name_embedding_list,
            threshold=THRESHOLD_EMBEDDINGS,
            limit=5
        )
    
    if not candidates:
        return None
    
    # STEP 3: Tiebreak with RapidFuzz
    best_match = None
    best_score = 0
    
    for candidate in candidates:
        # Cosine similarity (already calculated by Supabase)
        cosine_sim = candidate.get('similarity', candidate.get('combined_similarity', 0))
        
        # Text similarity with RapidFuzz (normalized)
        normalized_query = clean_text_for_matching(partner_name)
        normalized_candidate = clean_text_for_matching(candidate['name'])
        fuzz_name = fuzz.ratio(normalized_query, normalized_candidate) / 100
        
        # If there's an acronym, compare it too
        fuzz_acronym = 0
        if acronym and candidate.get('acronym'):
            normalized_query_acronym = clean_text_for_matching(acronym)
            normalized_candidate_acronym = clean_text_for_matching(candidate['acronym'])
            fuzz_acronym = fuzz.ratio(normalized_query_acronym, normalized_candidate_acronym) / 100
        
        # Combined final score
        combined_score = (
            COSINE_WEIGHT * cosine_sim +
            FUZZ_NAME_WEIGHT * fuzz_name +
            FUZZ_ACRONYM_WEIGHT * fuzz_acronym
        )
        
        if combined_score > best_score:
            best_score = combined_score
            best_match = {
                'clarisa_id': candidate['clarisa_id'],
                'name': candidate['name'],
                'acronym': candidate.get('acronym', ''),
                'website': candidate.get('website', ''),
                'countries': candidate.get('countries', []),
                'institution_type': candidate.get('institution_type', ''),
                'cosine_similarity': cosine_sim,
                'fuzz_name_score': fuzz_name,
                'fuzz_acronym_score': fuzz_acronym,
                'final_score': combined_score
            }
    
    return best_match if best_score >= THRESHOLD_FINAL else None


# ============================================================================
# MAIN PIPELINE
# ============================================================================


def process_partners_to_json(df):
    """
    Process a DataFrame with partners and return results as JSON-serializable dict
    
    Args:
        df: DataFrame with partners (same format as Excel input)
        
    Returns:
        dict: Results with partners data and statistics
    """
    logger.info(f"⚙️  CONFIGURATION:")
    logger.info(f"   - Threshold embeddings: {THRESHOLD_EMBEDDINGS}")
    logger.info(f"   - Threshold final:      {THRESHOLD_FINAL}")
    logger.info(f"   - Name/acronym weight:  {NAME_WEIGHT}/{ACRONYM_WEIGHT}")
    logger.info(f"   - Final score:          {COSINE_WEIGHT*100:.0f}% cosine + {FUZZ_NAME_WEIGHT*100:.0f}% fuzz_name + {FUZZ_ACRONYM_WEIGHT*100:.0f}% fuzz_acronym")
    
    # Process each row
    logger.info(f"🔄 Processing {len(df)} institutions...")
    
    partners_results = []
    stats = {
        'total': len(df),
        'matched': 0,
        'no_match': 0,
        'web_search_attempted': 0,
        'web_search_success': 0,
        'errors': 0,
        'excellent': 0,
        'good': 0,
        'fair': 0
    }
    
    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Searching matches"):
        # Read data from Excel
        partner_id = row.iloc[0] if len(row) > 0 and pd.notna(row.iloc[0]) else ''
        partner_name = row.iloc[1] if len(row) > 1 else None
        acronym = row.iloc[2] if len(row) > 2 else None
        website = row.iloc[3] if len(row) > 3 and pd.notna(row.iloc[3]) else None
        country = row.iloc[5] if len(row) > 5 and pd.notna(row.iloc[5]) else None
        
        partner_data = {
            'id': str(partner_id),
            'name': str(partner_name) if partner_name else '',
            'acronym': str(acronym) if acronym and str(acronym).lower() not in ['nan', 'none', ''] else '',
            'website': str(website) if website else '',
            'country': str(country) if country else '',
            'match_found': False,
            'clarisa_match': None,
            'web_search': None,
            'match_quality': 'no_match'
        }
        
        try:
            # Search for match
            match = search_institution_for_excel(partner_name, acronym)
            
            if match:
                # Match found
                score = match['final_score']
                if score >= 0.85:
                    quality = 'excellent'
                    stats['excellent'] += 1
                elif score >= 0.70:
                    quality = 'good'
                    stats['good'] += 1
                else:
                    quality = 'fair'
                    stats['fair'] += 1
                
                partner_data['match_found'] = True
                partner_data['match_quality'] = quality
                partner_data['clarisa_match'] = {
                    'clarisa_id': match['clarisa_id'],
                    'name': match['name'],
                    'acronym': match['acronym'] or '',
                    'countries': match['countries'],
                    'institution_type': match['institution_type'] or '',
                    'website': match['website'] or '',
                    'scores': {
                        'cosine_similarity': round(match['cosine_similarity'], 4),
                        'fuzz_name_score': round(match['fuzz_name_score'], 4),
                        'fuzz_acronym_score': round(match['fuzz_acronym_score'], 4),
                        'final_score': round(match['final_score'], 4)
                    }
                }
                
                stats['matched'] += 1
            else:
                # No match in CLARISA - Try web search if enabled
                if ENABLE_WEB_SEARCH:
                    stats['web_search_attempted'] += 1
                    web_result = search_institution_online(partner_name, country, website)
                    
                    if web_result['success']:
                        partner_data['web_search'] = {
                            'success': True,
                            'result': web_result.get('formatted_result', '')
                        }
                        stats['web_search_success'] += 1
                    else:
                        partner_data['web_search'] = {
                            'success': False,
                            'error': web_result.get('error', 'Unknown error')
                        }
                
                stats['no_match'] += 1
                
        except Exception as e:
            # Error processing
            logger.error(f"\n⚠️  Error in row {idx}: {e}")
            partner_data['match_quality'] = 'error'
            partner_data['error'] = str(e)
            stats['errors'] += 1
        
        partners_results.append(partner_data)
    
    # Calculate percentages
    stats['matched_percentage'] = round(stats['matched'] / stats['total'] * 100, 1) if stats['total'] > 0 else 0
    stats['no_match_percentage'] = round(stats['no_match'] / stats['total'] * 100, 1) if stats['total'] > 0 else 0
    
    return {
        'partners': partners_results,
        'stats': stats
    }


def run_pipeline(excel_path):
    """
    Main pipeline: processes an Excel file with institutions and maps them to CLARISA
    
    Args:
        excel_path: Path to Excel file
        
    Expected Excel format:
        - Column 0: ID (optional)
        - Column 1: partner_name
        - Column 2: acronym
        - Additional columns are preserved in the output
    """
    logger.info("=" * 80)
    logger.info("🚀 MAPPING PIPELINE: CGSpace → CLARISA")
    logger.info("=" * 80)
    
    # Verify that the DB has data
    logger.info("🔍 Verifying CLARISA database...")
    total_institutions = count_institutions()
    
    if total_institutions == 0:
        logger.error("❌ Error: The database is empty.")
        logger.error("   Run first: python populate_clarisa_db.py")
        return
    
    logger.info(f"✅ Database ready: {total_institutions:,} institutions in CLARISA")
    
    # Load Excel
    logger.info(f"📊 Loading Excel file: {excel_path}")
    try:
        df = pd.read_excel(excel_path)
        logger.info(f"✅ {len(df)} rows loaded")
    except Exception as e:
        logger.error(f"❌ Error loading Excel: {e}")
        return
    
    # Verify columns
    if len(df.columns) < 3:
        logger.error(f"❌ Error: Excel must have at least 3 columns")
        logger.error(f"   Format: [ID, Name, Acronym, ...]")
        logger.error(f"   Found: {len(df.columns)} columns")
        return
    
    logger.info("⚙️  CONFIGURATION:")
    logger.info(f"   - Threshold embeddings: {THRESHOLD_EMBEDDINGS}")
    logger.info(f"   - Threshold final:      {THRESHOLD_FINAL}")
    logger.info(f"   - Name/acronym weight:  {NAME_WEIGHT}/{ACRONYM_WEIGHT}")
    logger.info(f"   - Final score:          {COSINE_WEIGHT*100:.0f}% cosine + {FUZZ_NAME_WEIGHT*100:.0f}% fuzz_name + {FUZZ_ACRONYM_WEIGHT*100:.0f}% fuzz_acronym")
    
    # Process each row
    logger.info(f"🔄 Processing {len(df)} institutions...")
    
    results = {
        'match_found': [],
        'clarisa_id': [],
        'clarisa_name': [],
        'clarisa_acronym': [],
        'clarisa_countries': [],
        'clarisa_type': [],
        'clarisa_website': [],
        'cosine_similarity': [],
        'fuzz_name_score': [],
        'fuzz_acronym_score': [],
        'final_score': [],
        'match_quality': [],  # 'excellent', 'good', 'fair', 'no_match'
        'web_search_result': []  # Formatted web search result (single column)
    }
    
    stats = {
        'matched': 0,
        'no_match': 0,
        'web_search_attempted': 0,
        'web_search_success': 0,
        'errors': 0
    }
    
    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Searching matches"):
        # Read name and acronym from Excel
        # Assuming: column 1 = partner_name, column 2 = acronym
        partner_name = row.iloc[1] if len(row) > 1 else None
        acronym = row.iloc[2] if len(row) > 2 else None
        
        try:
            # Search for match
            match = search_institution_for_excel(partner_name, acronym)
            
            if match:
                # Match found
                results['match_found'].append(True)
                results['clarisa_id'].append(match['clarisa_id'])
                results['clarisa_name'].append(match['name'])
                results['clarisa_acronym'].append(match['acronym'] or '')
                results['clarisa_countries'].append(', '.join(match['countries']) if match['countries'] else '')
                results['clarisa_type'].append(match['institution_type'] or '')
                results['clarisa_website'].append(match['website'] or '')
                results['cosine_similarity'].append(round(match['cosine_similarity'], 4))
                results['fuzz_name_score'].append(round(match['fuzz_name_score'], 4))
                results['fuzz_acronym_score'].append(round(match['fuzz_acronym_score'], 4))
                results['final_score'].append(round(match['final_score'], 4))
                
                # Determine match quality
                score = match['final_score']
                if score >= 0.85:
                    quality = 'excellent'
                elif score >= 0.70:
                    quality = 'good'
                else:
                    quality = 'fair'
                results['match_quality'].append(quality)
                results['web_search_result'].append('')  # No web search needed when match found
                
                stats['matched'] += 1
            else:
                # No match in CLARISA - Try web search if enabled
                web_search_formatted_result = ''
                
                if ENABLE_WEB_SEARCH:
                    # Get country and website from Excel if available
                    country = None
                    website = None
                    
                    # Try to get country from column 5 (if exists)
                    if len(row) > 5 and pd.notna(row.iloc[5]):
                        country = str(row.iloc[5]).strip()
                    
                    # Try to get website from column 3 (if exists)
                    if len(row) > 3 and pd.notna(row.iloc[3]):
                        website = str(row.iloc[3]).strip()
                    
                    # Perform TWO-PHASE web search
                    stats['web_search_attempted'] += 1
                    web_result = search_institution_online(partner_name, country, website)
                    
                    if web_result['success']:
                        web_search_formatted_result = sanitize_for_excel(web_result.get('formatted_result', ''))
                        stats['web_search_success'] += 1
                    else:
                        web_search_formatted_result = sanitize_for_excel(f"❌ WEB SEARCH FAILED: {web_result.get('error', 'Unknown error')}")
                
                # No match
                results['match_found'].append(False)
                results['clarisa_id'].append('')
                results['clarisa_name'].append('')
                results['clarisa_acronym'].append('')
                results['clarisa_countries'].append('')
                results['clarisa_type'].append('')
                results['clarisa_website'].append('')
                results['cosine_similarity'].append(0)
                results['fuzz_name_score'].append(0)
                results['fuzz_acronym_score'].append(0)
                results['final_score'].append(0)
                results['match_quality'].append('no_match')
                results['web_search_result'].append(web_search_formatted_result)
                
                stats['no_match'] += 1
                
        except Exception as e:
            # Error processing
            logger.error(f"\n⚠️  Error in row {idx}: {e}")
            results['match_found'].append(False)
            results['clarisa_id'].append('ERROR')
            results['clarisa_name'].append(str(e))
            results['clarisa_acronym'].append('')
            results['clarisa_countries'].append('')
            results['clarisa_type'].append('')
            results['clarisa_website'].append('')
            results['cosine_similarity'].append(0)
            results['fuzz_name_score'].append(0)
            results['fuzz_acronym_score'].append(0)
            results['final_score'].append(0)
            results['match_quality'].append('error')
            results['web_search_result'].append('')
            
            stats['errors'] += 1
    
    # Add results to DataFrame
    df['MATCH_FOUND'] = results['match_found']
    df['CLARISA_ID'] = results['clarisa_id']
    df['CLARISA_NAME'] = results['clarisa_name']
    df['CLARISA_ACRONYM'] = results['clarisa_acronym']
    df['CLARISA_COUNTRIES'] = results['clarisa_countries']
    df['CLARISA_TYPE'] = results['clarisa_type']
    df['CLARISA_WEBSITE'] = results['clarisa_website']
    df['COSINE_SIMILARITY'] = results['cosine_similarity']
    df['FUZZ_NAME_SCORE'] = results['fuzz_name_score']
    df['FUZZ_ACRONYM_SCORE'] = results['fuzz_acronym_score']
    df['FINAL_SCORE'] = results['final_score']
    df['MATCH_QUALITY'] = results['match_quality']
    df['WEB_SEARCH_RESULT'] = results['web_search_result']
    
    # Save results
    output_file = "clarisa_mapping_results.xlsx"
    logger.info(f"💾 Saving results to {output_file}...")
    
    # Save with basic formatting
    df.to_excel(output_file, sheet_name='Results', index=False, engine='openpyxl')
    
    # Apply text wrapping to WEB_SEARCH_RESULT column for readability
    from openpyxl import load_workbook
    workbook = load_workbook(output_file)
    worksheet = workbook['Results']
    
    # Set column widths and enable text wrapping
    for idx, col in enumerate(df.columns, 1):
        col_letter = worksheet.cell(1, idx).column_letter
        if col == 'WEB_SEARCH_RESULT':
            worksheet.column_dimensions[col_letter].width = 80
            # Enable text wrapping for this column
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=idx)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    workbook.save(output_file)
    workbook.close()
    
    logger.info("=" * 80)
    logger.info("📊 RESULTS")
    logger.info("=" * 80)
    logger.info(f"✅ File saved: {output_file}")
    logger.info(f"📈 Statistics:")
    logger.info(f"   Total processed:      {len(df):,}")
    logger.info(f"   ✅ Successful matches: {stats['matched']:,} ({stats['matched']/len(df)*100:.1f}%)")
    logger.info(f"   ❌ No match:           {stats['no_match']:,} ({stats['no_match']/len(df)*100:.1f}%)")
    if ENABLE_WEB_SEARCH and stats['web_search_attempted'] > 0:
        logger.info(f"   🌐 Web search:         {stats['web_search_success']}/{stats['web_search_attempted']} successful")
    logger.info(f"   ⚠️  Errors:             {stats['errors']:,}")
    
    # Breakdown by quality
    if stats['matched'] > 0:
        excellent = sum(1 for q in results['match_quality'] if q == 'excellent')
        good = sum(1 for q in results['match_quality'] if q == 'good')
        fair = sum(1 for q in results['match_quality'] if q == 'fair')
        
        logger.info(f"🏆 Match quality:")
        logger.info(f"   Excellent (≥0.85): {excellent:,} ({excellent/stats['matched']*100:.1f}%)")
        logger.info(f"   Good (≥0.70):      {good:,} ({good/stats['matched']*100:.1f}%)")
        logger.info(f"   Fair (≥0.50):      {fair:,} ({fair/stats['matched']*100:.1f}%)")
    
    # Average scores
    if stats['matched'] > 0:
        matched_scores = [s for s, q in zip(results['final_score'], results['match_quality']) if q != 'no_match' and q != 'error']
        avg_score = sum(matched_scores) / len(matched_scores) if matched_scores else 0
        max_score = max(matched_scores) if matched_scores else 0
        min_score = min(matched_scores) if matched_scores else 0
        
        logger.info(f"📊 Similarity scores:")
        logger.info(f"   Average: {avg_score:.4f}")
        logger.info(f"   Maximum: {max_score:.4f}")
        logger.info(f"   Minimum: {min_score:.4f}")
    
    logger.info("=" * 80)
    logger.info("✅ PROCESS COMPLETED")
    logger.info("=" * 80)